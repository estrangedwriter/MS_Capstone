from openpyxl import load_workbook
import config
import matplotlib.pyplot as plt

##### Class declaration called "VOE_Emails".
class GatherData:

    #### __init__ function declares the proper excel file, caselog, and
    #### a list called self.information. This program will rely on nested lists.
    def __init__(self):
        self.wb = load_workbook("0 GENERIC SWE LOG 2020.xlsx")
        self.caselog = self.wb["Current Caselog"]
        self.information = []
        self.fields = config.fields
    ####
    #### This function gathers all the relevant information from the log and cleans the information as well.
    def gatherinfoVOE(self):
        try:
            for row in self.caselog[self.fields]:
                info = []
        #### Each row/cell of information will be put into a temporary list called "info".
                for cell in row:
                    cell = cell.value
                    info.append(cell)
        #### Cleans each temporary list to be put into the "self.information" list.
        #### if a case is qualified then it will be added to the self.information list. If not,
        #### 'continue' (Skip over it)
                if info[15] != 'draft':
                    continue
                else:
                    self.information.append(info)
                    del info[4:7]
                    del info[6:11]
        except:
            pass
    ####
    #### This function gathers all the relevant information from the log and cleans the information as well.
    def gatherinfoJDReqs(self):
        try:
            for row in self.caselog[self.fields]:
                info = []
                ####Each row/cell of information will be put into a temporary list called "info".
                for cell in row:
                    cell = cell.value
                    info.append(cell)

                #### if a case is qualified then it will be added to the self.information list. If not,
                #### 'continue' (Skip over it)
                if info[12] != 'draft':
                    continue
                else:
                    del info[7:9]
                    del info[8:10]
                    del info[9:10]
                    self.information.append(info)
        except:
            pass
    ####
    #### This function gathers all the relevant information from the log and cleans the information as well.
    def gatherinfoETAReview(self):
        try:
            for row in self.caselog[self.fields]:
                info = []
                ####Each row/cell of information will be put into a temporary list called "info".
                for cell in row:
                    cell = cell.value
                    info.append(cell)

                #### if a case is qualified then it will be added to the self.information list. If not,
                #### 'continue' (Skip over it)
                if info[24] != 'draft':
                    continue
                else:
                    del info[8:24]
                    del info[4:7]
                    self.information.append(info)
        except:
            pass
    ####
    #### This function gathers all the relevant information from the log and cleans the information as well.
    def gatherinfoPERMFiled(self):
        try:
            for row in self.caselog[self.fields]:
                info = []
                ####Each row/cell of information will be put into a temporary list called "info".
                for cell in row:
                    cell = cell.value
                    info.append(cell)

                #### if a case is qualified then it will be added to the self.information list. If not,
                #### 'continue' (Skip over it)
                if info[30] != 'draft':
                    continue
                else:
                    del info[15:30]
                    del info[9:14]
                    del info[4:7]
                    self.information.append(info)
        except:
            pass
    ####
    def printinformation(self):
        config.information = self.information
        print (config.information)
    ####

#### Class declaration for the analytics dashboard
class Analyticsdashboard:

    def __init__(self):
        self.wb = load_workbook("0 GENERIC SWE LOG 2020.xlsx")      ### this is the generic log name
        self.caselog = self.wb["Current Caselog"]                   ### this is the sheet name

        self.counteroptions = {'A': 0,'B': 0, 'C1': 0, 'C2': 0, 'C': 0} ### the counteroption dictionary will count up all the possible options in the log
        self.counter = 0                                                 #### counter integer will count the total number of cases
        self.confirmed = {'Confirmed': 0, 'Unconfirmed': 0}             ### the confirmed dictionary counts the number of confirmed or unconfirmed VOE cases
        self.labelspie = []
        self.labelsbar = []
        self.figurespie = []
        self.figuresbar = []

    ### this function gathers all of the relevant information
    def gatherthedata(self):
        try:
            #### counts up to 1000 entries in the log (will never reach that point so it will gather all the data points in the log)
            ### this function counts the number of each generic option
            for row in self.caselog["Q03:Q1000"]:
                for cell in row:
                    ### if there is no 'final option' marked in the generic log, then do nothing
                    if cell.value not in self.counteroptions:
                        pass
                    ### if there is a final option marked, then count it up
                    else:
                        self.counteroptions[cell.value] += 1
                        self.counter = self.counter + 1
            ### this function counts the total number of confirmed or unconfirmed VOE's cases
            for row in self.caselog["S03:S1000"]:
                for cell in row:
                    #### if VOE is confirmed then add 1 to confirmed count
                    if cell.value == 'Y':
                        self.confirmed['Confirmed'] += 1
                    ### if VOE not confirmed then do nothing
                    else:
                        pass
        except:
            pass
        ####option C is just the total number of 'C1' + 'C2' cases
        self.counteroptions['C'] = self.counteroptions['C1'] + self.counteroptions['C2']
        ####  unconfirmed cases is total cases minus confirmed cases
        self.confirmed['Unconfirmed'] = self.counter - self.confirmed['Confirmed']

        ###### values for the pie chart
        self.labelspie = list(self.counteroptions.keys())          ### this is the labels for the pie chart (option a, b, c, the keys for the counteroptions dictionary
        self.figurespie = list(self.counteroptions.values())      ### this is the numbers for the pie chart (option a, b, c, the values for the counteroptions dictionary
        del self.figurespie[2:4]         ### delete options C1 and C2
        del self.labelspie[2:4]          ### delete values for C1 and C2 we only want total number of cases for C

        ####### values for the bar chart
        self.labelsbar = list(self.confirmed.keys())         ### this is the labels for the bar chart (confirmed or unconfirmed cases)
        self.figuresbar = list(self.confirmed.values())      ### this is the values for the bar chart (number of confirmed or unconfirmed cases)

    ##### this function plots the charts for the analytics dashboard
    def charts(self):

        ###this function will allow for the pie chart to show actual values
        def make_autopct(figures):
            def my_autopct(pct):
                total = sum(figures)
                val = int(round(pct * total / 100.0))
                return '{p:.2f}%  ({v:d})'.format(p=pct, v=val)
            return my_autopct
        #################
        ####### this is the informative legend plotted on the chart
        label1 = ("Total number of qualified Generic cases: {count}\n"
                  "Total number of cases w/confirmed VOE's: {number1}\n"
                  "Total number of cases without confirmed VOE's: {number2}".format(count=self.counter,
                                                                                   number1 = self.confirmed['Confirmed'],
                                                                                   number2 = self.confirmed['Unconfirmed']))
        ########
        ###### the charts
        fig1, (ax1, ax2) = plt.subplots(1, 2)  ###
        fig1.suptitle("Analytics Dashboard")  ### name of the dashboard
        #### configuring the pie chart
        explode = (0.35, 0, 0.1)
        colors = ['#ff9999', '#66b3ff', '#99ff99']
        ###### the pie chart
        ax1.pie(self.figurespie, explode = explode, labels = self.labelspie, startangle = 90, autopct = make_autopct(self.figurespie),
                shadow = True, colors = colors)
        ax1.axis('equal')
        ax1.legend()
        #### the bar chart
        ax2.bar(self.labelsbar, self.figuresbar, align = 'center', width = 0.4, label = label1)
        ax2.legend(bbox_to_anchor = (1, -0.05), shadow=True, ncol = 2)
        ### plotting the charts and making sure it fits within the screen
        plt.tight_layout()
        plt.show()
        ##############

### the main functions for each feature of the analytics
### This function gathers all relevant information for drafting VOE emails
def mainloopVOE():
    VOE = GatherData()
    VOE.gatherinfoVOE()
    VOE.printinformation()
### This function gathers all relevant information for drafting JD&Requirements
def mainloopJDReqs():
    JDReqs = GatherData()
    JDReqs.gatherinfoJDReqs()
    JDReqs.printinformation()
### This function gathers all relevant information for drafting ETA 9089 EE Review Emails
def mainloopETAReview():
    ETAReview = GatherData()
    ETAReview.gatherinfoETAReview()
    ETAReview.printinformation()
### This function gathers all relevant information for drafting PERM Filed Emails
def mainloopPERMFiled():
    PERMF = GatherData()
    PERMF.gatherinfoPERMFiled()
    PERMF.printinformation()
### This function
def mainanalytics():
    analyze = Analyticsdashboard()
    analyze.gatherthedata()
    analyze.charts()
####################